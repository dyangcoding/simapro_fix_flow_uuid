import xlrd
import glob
from openpyxl import load_workbook
import os
import xlwings as xw 
import pandas as pd 
import csv

class MappedData(object):

    def __init__(self):
        self.file = ''
        self.sheet = ''
        self.process = ''
        self.old_flow = ''
        self.new_flow = ''

    def __eq__(self, other):
        if isinstance(other, self.__class__):
            return self.process == other.process and self.old_flow == other.old_flow
        return False

    def __ne__(self, other):
        return self.process != other.process or self.old_flow != other.old_flow
        
def get_mapping() -> dict:
    """
    return a dict which mapped process to referenced flow with the correct uuid
    """
    process_flow = {}
    for xls in get_mapping_files():
        if xls != 'post_processing_fixFlowUUIDS_4SimaANDGabi.xlsx':
            continue
        print('read mappings from %s' % xls)
        workbook = xlrd.open_workbook(xls)
        sheet = workbook.sheet_by_index(0)  # type: xlrd.sheet.Sheet
        for row in range(1, sheet.nrows):
            process_ref_index, flow_ref_index = 0, 1 
            process_ref = sheet.cell_value(row, process_ref_index)
            flow_ref = sheet.cell_value(row, flow_ref_index)
            process_flow[process_ref] = flow_ref
    return process_flow

def get_mapping_files() -> str:
    for f in glob.glob('*.xlsx'):
        yield f

def iter_files():
    mapping = get_mapping()
    for root, _, files in os.walk('simapro'):
        for f in files:
            print('folder {0} contains files {1}'.format(root, f))
            check_mapping(os.path.join(root, f), mapping)
           
def check_mapping(file_path: str, mapping: dict):
    """
    iterate through all generated xlsx files and check if there 
    are process with referenced flow which has wrong uuid

    format of original file has been changed after modify
    """
    wb = load_workbook(file_path)
    print('load file {}'.format(file_path))
    for sheet in wb.sheetnames:
        if not sheet.endswith('processes'):
            continue
        ws = wb[sheet]
        print('read sheet {}'.format(sheet))
        flow_section = False
        for row, _ in enumerate(ws.rows):
            cell_uuid = 'D' + str(row)
            cell_provider = 'K' + str(row)
            uuid = ws[cell_uuid]
            provider = ws[cell_provider]
            if flow_section:
                if uuid is None and provider is None:
                   flow_section = False
                   continue
                if provider not in mapping:
                    continue
                if uuid != mapping[provider]:
                    ws[cell_uuid] = mapping[provider]
                    print('{0} replaced with {1}'.format(uuid, mapping[provider]))
            if uuid == 'UUID' and provider == 'Provider UUID':
                flow_section = True
    wb.save(file_path)

def get_xls_files():
    for root, _, files in os.walk('simapro'):
        for f in files:
            yield os.path.abspath(os.path.join(root, f))

def iter():
    mapped_data = []
    mapping = get_mapping()
    for path in get_xls_files():
        file_path = r'%s' % path
        print('read process uuids from file {}'.format(file_path))
        check_mapping_xw(path, mapping, mapped_data)
    write_csv(mapped_data)

def write_csv(data: list):
    with open('mapped_flow_uuid.csv', mode='w', newline='') as f:
        writer = csv.writer(f, delimiter = ',')
        writer.writerow(['File', 'Sheet', 'Process', 'Old flow', 'New Flow'])
        for d in data:
            writer.writerow([d.file, d.sheet, d.process, d.old_flow, d.new_flow])

def check_mapping_xw(file_path: str, mapping: dict, mapped_data: list):
    wb = xw.Book(file_path)
    app = xw.apps.active
    print('load file {}'.format(file_path))
    for sheet in wb.sheets:
        if not sheet.name.endswith('processes'):
            continue
        sheet = wb.sheets[sheet.name]
        sheet.activate()
        df = pd.read_excel(file_path, sheet_name=sheet.name)
        print('row numbers {}'.format(len(df)))

        flow_section = False
        for row in range(1, len(df)):
            cell_uuid = 'D' + str(row)
            cell_provider = 'K' + str(row)
            uuid = xw.Range(cell_uuid).value
            provider = xw.Range(cell_provider).value
            if flow_section:
                if uuid is None and provider is None:
                   flow_section = False
                   continue
                if provider not in mapping:
                    continue
                if uuid != mapping[provider]:
                    xw.Range(cell_uuid).value = mapping[provider]
                    mapped = MappedData()
                    mapped.file = file_path
                    mapped.sheet = sheet.name
                    mapped.process = provider
                    mapped.old_flow = uuid
                    mapped.new_flow = mapping[provider]
                    if mapped not in mapped_data:
                        mapped_data.append(mapped)
            if uuid == 'Flow UUID' and provider == 'Process data set UUID':
                flow_section = True
    wb.save()
    # close excel instance after processing
    app.quit()

if __name__ == '__main__':
    iter()